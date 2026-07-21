#!/usr/bin/env python3
"""
STARK096 — ALDEA: Generate Final Sheets from Real Input
Reads real client data from input/ Excel files and generates individual sheets
per housing unit.  One output file is created per input file, named:
  {input_basename}_{timestamp}.xlsx  →  output_final/

Input format (CLIENTES sheet):
  Row 5  = headers
  Row 6+ = data rows
  Col 1  = Nº row
  Col 2  = NOMBRE         (Comprador 1)
  Col 3  = APELLIDO 1     (Comprador 1)
  Col 4  = APELLIDO 2     (Comprador 1)
  Col 5  = NOMBRE         (Comprador 2, may be empty)
  Col 6  = APELLIDO 1     (Comprador 2, may be empty)
  Col 7  = APELLIDO 2     (Comprador 2, may be empty)
  Col 8  = ESC
  Col 9  = PLANTA
  Col 10 = LETRA
  Col 17 = MÁX DISPONIBLE  (not used for pre-fill per user preference)
"""

import os
import re
import glob
import openpyxl

from generate_final_sheets import (
    load_mejoras_catalog,
    create_summary_sheet,
    create_constructor_summary_sheet,
    populate_constructor_summary,
    create_header_box,
    create_mejoras_table,
)
from utils import (
    create_output_dir,
    get_timestamp,
    print_success,
    print_error,
    print_info,
    print_warning,
)

# ---------------------------------------------------------------------------
# Honorific stripping
# ---------------------------------------------------------------------------
_HONORIFIC_RE = re.compile(
    r'^(D(?:ÑA|OÑA|ON|R[Aa]|NA)?\.?|SR[Aa]?\.?)\s+',
    re.IGNORECASE | re.UNICODE,
)


def strip_honorific(name):
    """Remove leading honorific title from a name string.

    Examples: "DÑA. NURIA PILAR" → "NURIA PILAR"
              "DON. SERGIO"       → "SERGIO"
              "D. JORGE "         → "JORGE"
    """
    if not name:
        return name
    return _HONORIFIC_RE.sub('', str(name).strip()).strip()


# ---------------------------------------------------------------------------
# Input reader
# ---------------------------------------------------------------------------

def read_real_input(filepath):
    """Read housing data from a real client input Excel file.

    Returns a list of vivienda dicts with the same keys used by
    generate_final_sheets.py so that all shared functions work unchanged.
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        viviendas = []
        for row_idx in range(6, ws.max_row + 1):
            row_num = ws.cell(row=row_idx, column=1).value
            if row_num is None:
                continue  # blank row

            esc    = ws.cell(row=row_idx, column=8).value
            planta = ws.cell(row=row_idx, column=9).value
            letra  = ws.cell(row=row_idx, column=10).value

            if esc is None or planta is None or letra is None:
                print_warning(f"Row {row_idx}: missing ESC/PLANTA/LETRA — skipped")
                continue

            try:
                piso = f"E{int(esc)} {int(planta)}{str(letra).strip()}"
            except (ValueError, TypeError):
                print_warning(f"Row {row_idx}: cannot parse esc={esc!r} planta={planta!r} — skipped")
                continue

            nombre1    = strip_honorific(ws.cell(row=row_idx, column=2).value)
            apellido1_1 = _clean(ws.cell(row=row_idx, column=3).value)
            apellido1_2 = _clean(ws.cell(row=row_idx, column=4).value)
            nombre2    = strip_honorific(ws.cell(row=row_idx, column=5).value)
            apellido2_1 = _clean(ws.cell(row=row_idx, column=6).value)
            apellido2_2 = _clean(ws.cell(row=row_idx, column=7).value)

            viviendas.append({
                "Comprador 1 - Nombre":     nombre1    or "",
                "Comprador 1 - Apellido 1": apellido1_1 or "",
                "Comprador 1 - Apellido 2": apellido1_2 or "",
                "Comprador 2 - Nombre":     nombre2,
                "Comprador 2 - Apellido 1": apellido2_1,
                "Comprador 2 - Apellido 2": apellido2_2,
                "Piso": piso,
            })

        wb.close()
        print_info(f"Read {len(viviendas)} housing units from {os.path.basename(filepath)}")
        return viviendas

    except Exception as e:
        print_error(f"Error reading {filepath}: {e}")
        return None


def _clean(val):
    """Strip whitespace from a cell value; return None if empty."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


# ---------------------------------------------------------------------------
# Main generator
# ---------------------------------------------------------------------------

def generate_real_sheets():
    """Process every .xlsx in input/ and write one output file per input."""

    print_info("Loading mejoras catalog...")
    catalog = load_mejoras_catalog()
    if not catalog:
        return 1

    input_dir = "input"
    input_files = sorted(glob.glob(os.path.join(input_dir, "*.xlsx")))

    if not input_files:
        print_error(f"No .xlsx files found in {input_dir}/")
        return 1

    print_info(f"Found {len(input_files)} input file(s): " +
               ", ".join(os.path.basename(f) for f in input_files))

    output_dir = "output_final"
    create_output_dir(output_dir)

    timestamp    = get_timestamp()
    success_count = 0

    for input_file in input_files:
        input_basename = os.path.splitext(os.path.basename(input_file))[0]
        output_file    = os.path.join(output_dir, f"{input_basename}_{timestamp}.xlsx")

        print_info(f"\n── Processing: {os.path.basename(input_file)}")

        viviendas = read_real_input(input_file)
        if not viviendas:
            print_error(f"No valid data in {input_file} — skipping")
            continue

        wb = openpyxl.Workbook()

        print_info("  Creating summary sheet...")
        create_summary_sheet(wb, viviendas)

        print_info("  Creating constructor summary sheet...")
        create_constructor_summary_sheet(wb)

        print_info(f"  Creating {len(viviendas)} individual sheets...")
        for idx, vivienda in enumerate(viviendas, start=1):
            piso       = vivienda.get("Piso", f"Vivienda_{idx}")
            piso_short = (piso
                          .replace("Escalera ", "E")
                          .replace(" - Planta ", " ")
                          .replace(" - Puerta ", ""))
            sheet_name = piso_short[:31]

            print_info(f"    Sheet {idx}/{len(viviendas)}: {sheet_name}")
            ws = wb.create_sheet(title=sheet_name)
            header_end_row = create_header_box(ws, vivienda)
            create_mejoras_table(ws, vivienda, catalog, header_end_row)

        print_info("  Populating constructor summary with direct-reference formulas...")
        populate_constructor_summary(wb, catalog, viviendas)

        try:
            wb.save(output_file)
            print_success(f"Excel saved: {output_file}  ({len(wb.sheetnames)} sheets)")
            success_count += 1
        except Exception as e:
            print_error(f"Error saving {output_file}: {e}")

    if success_count == len(input_files):
        print_info(f"\nAll {success_count} file(s) generated successfully.")
        return 0
    else:
        print_warning(f"\n{success_count}/{len(input_files)} file(s) generated successfully.")
        return 1


if __name__ == "__main__":
    exit(generate_real_sheets())
