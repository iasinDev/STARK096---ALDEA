#!/usr/bin/env python3
"""
STARK096 — ALDEA: Utility Functions
Common utilities for Excel generation
"""

import os
import yaml
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def load_config(config_file="config/settings.yaml"):
    """Load configuration from YAML file"""
    if not os.path.exists(config_file):
        print(f"⚠️  Warning: Config file not found: {config_file}")
        return {}
    
    with open(config_file, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def create_output_dir(output_dir="output"):
    """Create output directory if it doesn't exist"""
    os.makedirs(output_dir, exist_ok=True)
    return output_dir


def get_timestamp():
    """Get current timestamp in standard format"""
    return datetime.now().strftime('%Y%m%d_%H%M%S')


def get_header_style(config=None):
    """Get header cell style based on configuration"""
    if config and 'excel' in config:
        excel_config = config['excel']
        fill = PatternFill(
            start_color=excel_config.get('header_color', '366092'),
            end_color=excel_config.get('header_color', '366092'),
            fill_type="solid"
        )
        font = Font(
            bold=True,
            color=excel_config.get('header_font_color', 'FFFFFF'),
            size=excel_config.get('header_font_size', 12),
            name=excel_config.get('default_font', 'Calibri')
        )
    else:
        # Default style
        fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    
    alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    return {
        'fill': fill,
        'font': font,
        'alignment': alignment,
        'border': border
    }


def get_cell_border():
    """Get standard cell border"""
    return Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def format_currency(value, config=None):
    """Format value as currency based on configuration"""
    if config and 'formats' in config:
        formats = config['formats']
        symbol = formats.get('currency_symbol', '€')
        decimal_sep = formats.get('decimal_separator', ',')
        thousands_sep = formats.get('thousands_separator', '.')
        
        # Format the number
        formatted = f"{value:,.2f}".replace(',', 'TEMP').replace('.', decimal_sep).replace('TEMP', thousands_sep)
        return f"{formatted} {symbol}"
    else:
        return f"{value:,.2f} €"


def validate_housing_data(data, config=None):
    """Validate housing data against configuration rules"""
    errors = []
    
    # Required fields check
    if config and 'validation' in config:
        required_fields = config['validation'].get('required_fields', [])
        for field in required_fields:
            if field not in data or not data[field]:
                errors.append(f"Missing required field: {field}")
        
        # Price validation
        if config['validation'].get('validate_prices', True):
            min_price = config['validation'].get('min_price', 0)
            max_price = config['validation'].get('max_price', float('inf'))
            
            if 'precio' in data:
                price = data['precio']
                if not isinstance(price, (int, float)):
                    errors.append(f"Invalid price type: {price}")
                elif price < min_price or price > max_price:
                    errors.append(f"Price out of range: {price} (allowed: {min_price}-{max_price})")
    
    return errors


def apply_header_style(cell, config=None):
    """Apply header style to a cell"""
    style = get_header_style(config)
    cell.fill = style['fill']
    cell.font = style['font']
    cell.alignment = style['alignment']
    cell.border = style['border']


def apply_cell_border(cell):
    """Apply border to a cell"""
    cell.border = get_cell_border()


def print_success(message):
    """Print success message with emoji"""
    print(f"✅ {message}")


def print_error(message):
    """Print error message with emoji"""
    print(f"❌ {message}")


def print_warning(message):
    """Print warning message with emoji"""
    print(f"⚠️  {message}")


def print_info(message):
    """Print info message with emoji"""
    print(f"ℹ️  {message}")


if __name__ == "__main__":
    # Test utilities
    print("🔧 Testing utility functions...")
    
    # Test config loading
    config = load_config()
    if config:
        print_success("Configuration loaded")
    
    # Test timestamp
    ts = get_timestamp()
    print_info(f"Current timestamp: {ts}")
    
    # Test currency formatting
    price = 185000
    formatted = format_currency(price, config)
    print_info(f"Formatted price: {formatted}")
    
    print_success("All utilities working correctly!")
