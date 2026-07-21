#!/usr/bin/env python3
"""
STARK096 — ALDEA: Generate Final Sheets
Creates individual sheets per housing unit from the base template with mejoras catalog
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.properties import PageSetupProperties
import os
import glob
import json
from utils import (
    create_output_dir,
    get_timestamp,
    print_success,
    print_error,
    print_info,
    print_warning
)


def load_mejoras_catalog():
    """Load the mejoras catalog from JSON file"""
    try:
        with open("mejoras_catalog.json", "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print_error(f"Error loading catalog: {e}")
        return None


def get_latest_template():
    """Get the most recent template file from output directory"""
    output_dir = "output"
    pattern = os.path.join(output_dir, "template_viviendas_*.xlsx")
    files = glob.glob(pattern)
    
    if not files:
        print_error("No template files found in output/ directory")
        return None
    
    latest_file = max(files, key=os.path.getmtime)
    print_info(f"Using template: {os.path.basename(latest_file)}")
    return latest_file


def read_housing_data(filepath):
    """Read housing data from the template Excel file"""
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            headers.append(header)
        
        viviendas = []
        for row in range(2, ws.max_row + 1):
            vivienda = {}
            for col in range(1, ws.max_column + 1):
                header = headers[col - 1]
                value = ws.cell(row=row, column=col).value
                vivienda[header] = value
            viviendas.append(vivienda)
        
        wb.close()
        print_info(f"Read {len(viviendas)} housing units")
        return viviendas
    
    except Exception as e:
        print_error(f"Error reading template: {e}")
        return None


def create_cuadro_mandos_sheet(wb):
    """Create the Cuadro de Mandos dashboard as the first sheet (uses wb.active)."""
    ws = wb.active
    ws.title = "Cuadro de mandos"

    title_fill    = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_font    = Font(bold=True, size=16, name="Calibri", color="FFFFFF")
    section_fill  = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    section_font  = Font(bold=True, size=11, name="Calibri", color="FFFFFF")
    button_fill   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    button_font   = Font(bold=True, size=13, name="Calibri", color="FFFFFF")
    button_border = Border(
        left=Side(style='medium', color='1F4E79'),
        right=Side(style='medium', color='1F4E79'),
        top=Side(style='medium', color='1F4E79'),
        bottom=Side(style='medium', color='1F4E79'),
    )
    note_font  = Font(italic=True, size=9, name="Calibri", color="595959")
    code_font  = Font(name="Courier New", size=8, color="595959")
    center_al  = Alignment(horizontal="center", vertical="center")
    left_al    = Alignment(horizontal="left",   vertical="center")

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 5
    for _cl in ['C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[_cl].width = 22
    ws.column_dimensions['H'].width = 3

    ws.row_dimensions[1].height  = 42
    ws.row_dimensions[2].height  = 10
    ws.row_dimensions[3].height  = 22
    ws.row_dimensions[4].height  = 8
    ws.row_dimensions[5].height  = 28
    ws.row_dimensions[6].height  = 28
    ws.row_dimensions[7].height  = 28
    ws.row_dimensions[8].height  = 12
    ws.row_dimensions[9].height  = 18
    ws.row_dimensions[10].height = 180

    def _fill_cols(row, col_start, col_end, fill, border=None):
        for _c in range(col_start, col_end + 1):
            _cell = ws.cell(row=row, column=_c)
            _cell.fill = fill
            if border:
                _cell.border = border

    # Row 1: main title
    ws.merge_cells('B1:G1')
    c = ws['B1']
    c.value = "\U0001f4ca  CUADRO DE MANDOS"
    c.font = title_font; c.fill = title_fill; c.alignment = center_al
    _fill_cols(1, 3, 7, title_fill)

    # Row 3: section header
    ws.merge_cells('B3:G3')
    c = ws['B3']
    c.value = "   \U0001f5a8\ufe0f  IMPRESI\u00d3N"
    c.font = section_font; c.fill = section_fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    _fill_cols(3, 3, 7, section_fill)

    # Rows 5-7: print button
    ws.merge_cells('B5:G7')
    c = ws['B5']
    c.value = "\U0001f5a8\ufe0f   Imprimir fichas de cliente"
    c.font = button_font; c.fill = button_fill
    c.alignment = center_al; c.border = button_border
    for _r in [5, 6, 7]:
        _fill_cols(_r, 3, 7, button_fill, button_border)

    # Row 9: usage note
    ws.merge_cells('B9:G9')
    c = ws['B9']
    c.value = ("\u2139\ufe0f  Para activar: Alt+F11 \u2192 Insertar m\u00f3dulo \u2192 pegar c\u00f3digo "
               "\u2192 clic derecho bot\u00f3n \u2192 Asignar macro \u2192 ImprimirFichasCliente")
    c.font = note_font; c.alignment = left_al

    # Row 10: VBA code block (single tall row for easy copy-paste)
    ws.merge_cells('B10:G10')
    _vba = (
        "Sub ImprimirFichasCliente()\n"
        "    Dim excluidas As Variant\n"
        "    excluidas = Array(\"Cuadro de mandos\", \"Resumen\", \"Resumen Constructora\")\n"
        "    Dim hojas() As String: Dim n As Integer: n = 0\n"
        "    Dim ws As Worksheet\n"
        "    For Each ws In ThisWorkbook.Sheets\n"
        "        Dim skip As Boolean: skip = False\n"
        "        Dim i As Integer\n"
        "        For i = 0 To UBound(excluidas)\n"
        "            If ws.Name = excluidas(i) Then skip = True: Exit For\n"
        "        Next i\n"
        "        If Not skip Then ReDim Preserve hojas(n): hojas(n) = ws.Name: n = n + 1\n"
        "    Next ws\n"
        "    If n = 0 Then MsgBox \"No hay fichas de cliente.\": Exit Sub\n"
        "    ThisWorkbook.Sheets(hojas).Select\n"
        "    ActiveWindow.SelectedSheets.PrintOut Copies:=1\n"
        "    ThisWorkbook.Sheets(\"Cuadro de mandos\").Select\n"
        "    MsgBox n & \" fichas enviadas a impresora.\", vbInformation\n"
        "End Sub"
    )
    c = ws['B10']
    c.value = _vba
    c.font = code_font
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def create_summary_sheet(wb, viviendas):
    """Create the summary sheet with all viviendas - estilo a.png"""
    ws = wb.create_sheet(title="Resumen", index=1)
    
    # Define styles
    header_font = Font(bold=True, size=11, name="Calibri", color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Title row
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = "Importes Viviendas Totales"
    title_cell.font = Font(bold=True, size=14, name="Calibri")
    title_cell.alignment = center_align
    
    # Use "MÁX DISPONIBLE" header if viviendas carry pre-filled values, else "MAX MÓDULO"
    has_max_disponible = any(v.get("Max Disponible") is not None for v in viviendas)
    max_modulo_header = "MÁX DISPONIBLE" if has_max_disponible else "MAX MÓDULO"

    # Headers row 2
    headers = [
        "COMPRADOR 1 NOMBRE",
        "COMPRADOR 1 APELLIDO 1",
        "COMPRADOR 1 APELLIDO 2",
        "COMPRADOR 2 NOMBRE",
        "COMPRADOR 2 APELLIDO 1",
        "COMPRADOR 2 APELLIDO 2",
        "PISO",
        "TOTAL GASTADO",
        max_modulo_header,
        "SOBRANTE"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = center_align
    
    # Data rows starting from row 3
    for row_idx, vivienda in enumerate(viviendas, start=3):
        piso_short = vivienda.get("Piso", "").replace("Escalera ", "E").replace(" - Planta ", " ").replace(" - Puerta ", "")
        
        # Comprador 1
        ws.cell(row=row_idx, column=1, value=vivienda.get("Comprador 1 - Nombre")).border = border_style
        ws.cell(row=row_idx, column=2, value=vivienda.get("Comprador 1 - Apellido 1")).border = border_style
        ws.cell(row=row_idx, column=3, value=vivienda.get("Comprador 1 - Apellido 2")).border = border_style
        
        # Comprador 2
        ws.cell(row=row_idx, column=4, value=vivienda.get("Comprador 2 - Nombre")).border = border_style
        ws.cell(row=row_idx, column=5, value=vivienda.get("Comprador 2 - Apellido 1")).border = border_style
        ws.cell(row=row_idx, column=6, value=vivienda.get("Comprador 2 - Apellido 2")).border = border_style
        
        # Piso
        ws.cell(row=row_idx, column=7, value=piso_short).border = border_style
        
        # Total Gastado - formula referencing individual sheet (C7 = TOTAL GASTADO value, row 7 is fixed)
        cell_total = ws.cell(row=row_idx, column=8)
        cell_total.value = f"='{piso_short}'!C7"
        cell_total.number_format = '#,##0.00'
        cell_total.border = border_style
        
        # Max Disponible / Max Módulo — pre-filled from input when available, else empty
        max_disp_val = vivienda.get("Max Disponible")
        cell_max = ws.cell(row=row_idx, column=9, value=max_disp_val if max_disp_val is not None else "")
        cell_max.border = border_style
        cell_max.number_format = '#,##0.00'
        
        # Sobrante - formula =MAX MÓDULO - TOTAL GASTADO
        cell_sobrante = ws.cell(row=row_idx, column=10)
        cell_sobrante.value = f"=I{row_idx}-H{row_idx}"
        cell_sobrante.number_format = '#,##0.00'
        cell_sobrante.border = border_style
    
    # Conditional formatting: highlight TOTAL GASTADO in pastel red when it exceeds MÁX DISPONIBLE
    if has_max_disponible and viviendas:
        last_data_row = len(viviendas) + 2  # data starts at row 3
        pastel_red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        # Formula anchored to first data row; Excel adjusts per row automatically.
        # Only fires when col I has a value (real mode) AND col H exceeds it.
        rule = FormulaRule(formula=['=AND(I3<>"",H3>I3)'], fill=pastel_red)
        ws.conditional_formatting.add(f"H3:H{last_data_row}", rule)

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15


def create_constructor_summary_sheet(wb):
    """Create constructor summary sheet with aggregated items (headers only)"""
    ws = wb.create_sheet(title="Resumen Constructora", index=2)
    
    # Define styles
    header_font = Font(bold=True, size=11, name="Calibri", color="FFFFFF")
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Title
    ws.merge_cells('A1:C1')
    cell_title = ws.cell(row=1, column=1, value="LISTADO DE OBRA - RESUMEN CONSTRUCTORA")
    cell_title.font = Font(bold=True, size=14, name="Calibri")
    cell_title.alignment = center_align
    
    # Headers
    headers = ["CÓDIGO", "CONCEPTO", "CANTIDAD TOTAL"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = center_align
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 18


def populate_constructor_summary(wb, catalog, viviendas):
    """Populate constructor summary with direct cell references to aggregate mejoras from all sheets.
    Uses F-column direct references instead of SUMIFS to stay within Excel's 8192-char formula limit.
    
    Layout contract (must match create_header_box + create_mejoras_table):
      - Header always occupies rows 1-7 (TOTAL GASTADO at row 7)
      - Table headers at row 9, data starts at row 10
      - Pasos basicos: rows 10 to 10+N_PASOS-1
      - Mejoras: start at row 10+N_PASOS, one row per mejora in catalog order
    """
    ws = wb["Resumen Constructora"]
    
    # Define styles
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    print_info("Creating direct-reference formulas for constructor summary...")
    
    # Collect all sheet names (same transformation used when creating sheets)
    sheet_names = []
    for vivienda in viviendas:
        piso = vivienda.get("Piso", "")
        piso_short = piso.replace("Escalera ", "E").replace(" - Planta ", " ").replace(" - Puerta ", "")
        sheet_names.append(piso_short[:31])
    
    # Compute the row where mejoras start in every individual sheet
    # Header always = 7 rows; table header offset = +2; data offset = +3
    DATA_START_ROW = 7 + 3  # = 10
    N_PASOS = len(catalog["pasos_basicos"])  # always same (e.g. 5)
    MEJORAS_START_ROW = DATA_START_ROW + N_PASOS  # e.g. 15
    
    current_row = 4
    mejora_idx = 0
    
    for mejora_cat in catalog.get("mejoras", []):
        for item in mejora_cat.get("items", []):
            concepto = item.get("concepto", "")
            if not concepto:
                continue
            
            codigo = f"MEJORA-{mejora_idx + 1:03d}"
            item_row = MEJORAS_START_ROW + mejora_idx
            
            # Column A: Código
            ws.cell(row=current_row, column=1, value=codigo).border = border_style
            ws.cell(row=current_row, column=1).alignment = center_align
            
            # Column B: Concepto
            ws.cell(row=current_row, column=2, value=concepto).border = border_style
            ws.cell(row=current_row, column=2).alignment = left_align
            
            # Column C: Direct sum of F{item_row} across all individual sheets
            # Formula: ='E1 1A'!F15+'E1 1B'!F15+... (much shorter than SUMIFS)
            formula = "=" + "+".join([f"'{sn}'!F{item_row}" for sn in sheet_names])
            cell_qty = ws.cell(row=current_row, column=3)
            cell_qty.value = formula
            cell_qty.border = border_style
            cell_qty.alignment = center_align
            
            mejora_idx += 1
            current_row += 1
    
    print_info(f"Created {mejora_idx} direct-reference formula rows (mejoras only)")


def create_header_box(ws, vivienda, origin_header=None):
    """Create the header box — FICHA DE CLIENTE.

    Layout (columns A–F, always 7 rows so TOTAL GASTADO is always at C7):
      Row 1: FICHA DE CLIENTE (A:C) | origin info (D:F) — or A:F if no origin
      Row 2: label (A:B) | comprador 1 full name (C:F)
      Row 3: empty label (A:B) | comprador 2 name or empty (C:F)  [always present]
      Row 4: separator (A:F)
      Row 5: Vivienda label (A:B) | piso value (C:F)
      Row 6: empty full-width (A:F)
      Row 7: TOTAL GASTADO label (A:B) | value cell (C:D) | € + IVA (E:F)
    """
    title_fill   = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    title_font   = Font(bold=True, size=12, name="Calibri", color="FFFFFF")
    origin_font  = Font(italic=True, size=9,  name="Calibri", color="EEEEEE")
    label_font   = Font(bold=True, size=11, name="Calibri", color="000000")
    value_font   = Font(size=11, name="Calibri", color="000000")
    total_fill   = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    total_font   = Font(bold=True, size=11, name="Calibri", color="000000")
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")
    right_align  = Alignment(horizontal="right",  vertical="center")

    def _brow(row, col_start, col_end, fill=None):
        """Apply border (and optional fill) to cols col_start..col_end in row."""
        for _c in range(col_start, col_end + 1):
            _cell = ws.cell(row=row, column=_c)
            _cell.border = border_style
            if fill:
                _cell.fill = fill

    current_row = 1

    # ── Row 1: FICHA DE CLIENTE + optional origin ──────────────────────────
    if origin_header:
        ws.merge_cells(f'A{current_row}:C{current_row}')
        cell = ws.cell(row=current_row, column=1, value="FICHA DE CLIENTE")
        cell.font = title_font; cell.fill = title_fill
        cell.border = border_style; cell.alignment = center_align
        _brow(current_row, 2, 3, title_fill)

        ws.merge_cells(f'D{current_row}:F{current_row}')
        cell_o = ws.cell(row=current_row, column=4, value=origin_header)
        cell_o.font = origin_font; cell_o.fill = title_fill
        cell_o.border = border_style
        cell_o.alignment = Alignment(horizontal="right", vertical="center")
        _brow(current_row, 5, 6, title_fill)
    else:
        ws.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws.cell(row=current_row, column=1, value="FICHA DE CLIENTE")
        cell.font = title_font; cell.fill = title_fill
        cell.border = border_style; cell.alignment = center_align
        _brow(current_row, 2, 6, title_fill)
    current_row += 1

    # ── Row 2: Comprador 1 ────────────────────────────────────────────────
    comprador1_nombre = vivienda.get("Comprador 1 - Nombre")
    comprador1_ap1    = vivienda.get("Comprador 1 - Apellido 1")
    comprador1_ap2    = vivienda.get("Comprador 1 - Apellido 2")
    compradores_text  = (
        f"{comprador1_nombre} {comprador1_ap1} {comprador1_ap2}"
        if comprador1_nombre and comprador1_ap1 and comprador1_ap2 else ""
    )
    comprador2_nombre = vivienda.get("Comprador 2 - Nombre")
    comprador2_ap1    = vivienda.get("Comprador 2 - Apellido 1")
    comprador2_ap2    = vivienda.get("Comprador 2 - Apellido 2")
    has_second_buyer  = comprador2_nombre and comprador2_ap1 and comprador2_ap2

    ws.merge_cells(f'A{current_row}:B{current_row}')
    cl = ws.cell(row=current_row, column=1, value="Nombre y Apellidos:")
    cl.font = label_font; cl.border = border_style; cl.alignment = right_align
    ws.cell(row=current_row, column=2).border = border_style

    ws.merge_cells(f'C{current_row}:F{current_row}')
    cv = ws.cell(row=current_row, column=3, value=compradores_text)
    cv.font = value_font; cv.border = border_style; cv.alignment = left_align
    _brow(current_row, 4, 6)
    current_row += 1

    # ── Row 3: Comprador 2 (always present for fixed-row layout) ─────────
    comprador2_text = (
        f"{comprador2_nombre} {comprador2_ap1} {comprador2_ap2}"
        if has_second_buyer else ""
    )
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws.cell(row=current_row, column=1).border = border_style
    ws.cell(row=current_row, column=2).border = border_style

    ws.merge_cells(f'C{current_row}:F{current_row}')
    cv = ws.cell(row=current_row, column=3, value=comprador2_text)
    cv.font = value_font; cv.border = border_style; cv.alignment = left_align
    _brow(current_row, 4, 6)
    current_row += 1

    # ── Row 4: Separator ─────────────────────────────────────────────────
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws.cell(row=current_row, column=1).border = border_style
    _brow(current_row, 2, 6)
    current_row += 1

    # ── Row 5: Vivienda ──────────────────────────────────────────────────
    piso           = vivienda.get("Piso", "")
    vivienda_short = (piso.replace("Escalera ", "E")
                         .replace(" - Planta ", " ")
                         .replace(" - Puerta ", ""))
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cl = ws.cell(row=current_row, column=1, value="Vivienda:")
    cl.font = label_font; cl.border = border_style; cl.alignment = right_align
    ws.cell(row=current_row, column=2).border = border_style

    ws.merge_cells(f'C{current_row}:F{current_row}')
    cv = ws.cell(row=current_row, column=3, value=vivienda_short)
    cv.font = value_font; cv.border = border_style; cv.alignment = left_align
    _brow(current_row, 4, 6)
    current_row += 1

    # ── Row 6: Empty full-width ──────────────────────────────────────────
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws.cell(row=current_row, column=1).border = border_style
    _brow(current_row, 2, 6)
    current_row += 1

    # ── Row 7: TOTAL GASTADO (value always at C7 — formula reference target)
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cl = ws.cell(row=current_row, column=1, value="TOTAL GASTADO:")
    cl.font = total_font; cl.fill = total_fill
    cl.border = border_style; cl.alignment = right_align
    ws.cell(row=current_row, column=2).fill   = total_fill
    ws.cell(row=current_row, column=2).border = border_style

    # C:D merged — value cell (formula written later by create_mejoras_table at C{row})
    ws.merge_cells(f'C{current_row}:D{current_row}')
    cv = ws.cell(row=current_row, column=3, value="")
    cv.fill = total_fill; cv.border = border_style
    cv.alignment = right_align; cv.number_format = '#,##0.00'
    ws.cell(row=current_row, column=4).fill   = total_fill
    ws.cell(row=current_row, column=4).border = border_style

    # E:F merged — currency label
    ws.merge_cells(f'E{current_row}:F{current_row}')
    cc = ws.cell(row=current_row, column=5, value="€ + IVA")
    cc.font = value_font; cc.fill = total_fill
    cc.border = border_style; cc.alignment = left_align
    ws.cell(row=current_row, column=6).fill   = total_fill
    ws.cell(row=current_row, column=6).border = border_style

    # Column widths (overridden by create_mejoras_table; kept as A4-friendly fallback)
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12

    return current_row


def _apply_a4_print_setup(ws):
    """Configure a worksheet for A4 portrait printing that fits to 1 page wide."""
    if not ws.sheet_properties.pageSetUpPr:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.paperSize   = 9           # A4
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0           # unlimited pages tall
    ws.print_options.horizontalCentered = True


def get_cocina_options_for_vivienda(piso_short, catalog):
    """Get cocina options for specific vivienda"""
    cocina_paso = None
    for paso in catalog["pasos_basicos"]:
        if paso["codigo"] == "COCINA":
            cocina_paso = paso
            break
    
    if not cocina_paso:
        return []
    
    for grupo in cocina_paso["grupos_por_tipo"]:
        if piso_short in grupo["tipos"]:
            return grupo["opciones"]
    
    return []


def create_mejoras_table(ws, vivienda, catalog, start_row):
    """Create mejoras selection table - estilo b.png"""
    
    header_fill = PatternFill(start_color="C6AC45", end_color="C6AC45", fill_type="solid")
    header_font = Font(bold=True, size=10, name="Calibri", color="000000")
    
    total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    total_font = Font(bold=True, size=11, name="Calibri")
    
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    current_row = start_row + 2
    
    # Headers
    headers = ["Concepto", "Selección", "Precio (€)", "Cantidad", "Seleccionado (Sí/No)", "Importe (€)"]
    cols = ['A', 'B', 'C', 'D', 'E', 'F']
    
    for col, header in zip(cols, headers):
        cell = ws[f'{col}{current_row}']
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = center_align
    
    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12
    
    current_row += 1
    data_start_row = current_row
    
    piso_short = vivienda.get("Piso", "").replace("Escalera ", "E").replace(" - Planta ", " ").replace(" - Puerta ", "")
    
    # Add basic pasos (always add COCINA row for consistent layout across all sheets)
    for paso in catalog["pasos_basicos"]:
        if paso["codigo"] == "COCINA":
            opciones = get_cocina_options_for_vivienda(piso_short, catalog)
            ws[f'A{current_row}'].value = paso["nombre"]
            ws[f'A{current_row}'].border = border_style
            ws[f'A{current_row}'].alignment = left_align
            
            ws[f'B{current_row}'].border = border_style
            if opciones:
                opciones_str = ",".join([opt["nombre"] for opt in opciones])
                dv = DataValidation(type="list", formula1=f'"{opciones_str}"', allow_blank=True)
                ws.add_data_validation(dv)
                dv.add(f'B{current_row}')
            
            ws[f'C{current_row}'].value = 0
            ws[f'C{current_row}'].border = border_style
            ws[f'C{current_row}'].alignment = right_align
            ws[f'C{current_row}'].number_format = '#,##0.00'
            
            ws[f'D{current_row}'].value = 0
            ws[f'D{current_row}'].border = border_style
            ws[f'D{current_row}'].alignment = center_align
            
            ws[f'E{current_row}'].value = "No"
            ws[f'E{current_row}'].border = border_style
            if opciones:
                dv_sino = DataValidation(type="list", formula1='"Sí,No"', allow_blank=True)
                ws.add_data_validation(dv_sino)
                dv_sino.add(f'E{current_row}')
            
            ws[f'F{current_row}'].value = f'=IF(E{current_row}="Sí",C{current_row}*D{current_row},0)'
            ws[f'F{current_row}'].border = border_style
            ws[f'F{current_row}'].alignment = right_align
            ws[f'F{current_row}'].number_format = '#,##0.00'
            
            current_row += 1
        else:
            ws[f'A{current_row}'].value = paso["nombre"]
            ws[f'A{current_row}'].border = border_style
            ws[f'A{current_row}'].alignment = left_align
            
            ws[f'B{current_row}'].border = border_style
            opciones_str = ",".join([opt["nombre"] for opt in paso["opciones"]])
            dv = DataValidation(type="list", formula1=f'"{opciones_str}"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f'B{current_row}')
            
            ws[f'C{current_row}'].value = 0
            ws[f'C{current_row}'].border = border_style
            ws[f'C{current_row}'].alignment = right_align
            ws[f'C{current_row}'].number_format = '#,##0.00'
            
            ws[f'D{current_row}'].value = 0
            ws[f'D{current_row}'].border = border_style
            ws[f'D{current_row}'].alignment = center_align
            
            ws[f'E{current_row}'].value = "No"
            ws[f'E{current_row}'].border = border_style
            dv_sino = DataValidation(type="list", formula1='"Sí,No"', allow_blank=True)
            ws.add_data_validation(dv_sino)
            dv_sino.add(f'E{current_row}')
            
            ws[f'F{current_row}'].value = f'=IF(E{current_row}="Sí",C{current_row}*D{current_row},0)'
            ws[f'F{current_row}'].border = border_style
            ws[f'F{current_row}'].alignment = right_align
            ws[f'F{current_row}'].number_format = '#,##0.00'
            
            current_row += 1
    
    # Add ALL mejoras from catalog
    if "mejoras" in catalog:
        for mejora_cat in catalog["mejoras"]:
            if "items" in mejora_cat:
                for item in mejora_cat["items"]:
                    concepto = item.get("concepto", "")
                    precio = item.get("precio", 0)
                    
                    if concepto:
                        ws[f'A{current_row}'].value = concepto
                        ws[f'A{current_row}'].border = border_style
                        ws[f'A{current_row}'].alignment = left_align
                        
                        ws[f'B{current_row}'].value = ""
                        ws[f'B{current_row}'].border = border_style
                        
                        ws[f'C{current_row}'].value = precio
                        ws[f'C{current_row}'].border = border_style
                        ws[f'C{current_row}'].alignment = right_align
                        ws[f'C{current_row}'].number_format = '#,##0.00'
                        
                        ws[f'D{current_row}'].value = 0
                        ws[f'D{current_row}'].border = border_style
                        ws[f'D{current_row}'].alignment = center_align
                        
                        ws[f'E{current_row}'].value = "No"
                        ws[f'E{current_row}'].border = border_style
                        dv_sino = DataValidation(type="list", formula1='"Sí,No"', allow_blank=True)
                        ws.add_data_validation(dv_sino)
                        dv_sino.add(f'E{current_row}')
                        
                        ws[f'F{current_row}'].value = f'=IF(E{current_row}="Sí",C{current_row}*D{current_row},0)'
                        ws[f'F{current_row}'].border = border_style
                        ws[f'F{current_row}'].alignment = right_align
                        ws[f'F{current_row}'].number_format = '#,##0.00'
                        
                        current_row += 1
    
    data_end_row = current_row - 1
    
    # TOTAL row
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:E{current_row}')
    cell_total_label = ws[f'A{current_row}']
    cell_total_label.value = "TOTAL"
    cell_total_label.font = total_font
    cell_total_label.fill = total_fill
    cell_total_label.border = border_style
    cell_total_label.alignment = right_align
    
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{current_row}'].border = border_style
        ws[f'{col}{current_row}'].fill = total_fill
    
    cell_total_value = ws[f'F{current_row}']
    cell_total_value.value = f'=SUM(F{data_start_row}:F{data_end_row})'
    cell_total_value.font = total_font
    cell_total_value.fill = total_fill
    cell_total_value.border = border_style
    cell_total_value.alignment = right_align
    cell_total_value.number_format = '#,##0.00'
    
    # Update TOTAL GASTADO in header (column C of the header row)
    ws[f'C{start_row}'].value = f'=F{current_row}'
    
    return current_row


def generate_final_sheets():
    """Generate the final Excel with individual sheets"""
    
    print_info("Loading mejoras catalog...")
    catalog = load_mejoras_catalog()
    if not catalog:
        return 1
    
    print_info("Generating final Excel...")
    
    template_file = get_latest_template()
    if not template_file:
        return 1
    
    viviendas = read_housing_data(template_file)
    if not viviendas:
        return 1
    
    output_dir = "output_final"
    create_output_dir(output_dir)
    
    wb = openpyxl.Workbook()

    print_info("Creating Cuadro de mandos...")
    create_cuadro_mandos_sheet(wb)

    print_info("Creating summary sheet...")
    create_summary_sheet(wb, viviendas)

    print_info("Creating constructor summary sheet...")
    create_constructor_summary_sheet(wb)

    print_info(f"Creating {len(viviendas)} individual sheets...")

    for idx, vivienda in enumerate(viviendas, start=1):
        piso = vivienda.get("Piso", f"Vivienda_{idx}")
        piso_short = piso.replace("Escalera ", "E").replace(" - Planta ", " ").replace(" - Puerta ", "")

        sheet_name = piso_short[:31]

        print_info(f"  Sheet {idx}/{len(viviendas)}: {sheet_name}")

        ws = wb.create_sheet(title=sheet_name)

        header_end_row = create_header_box(ws, vivienda)
        create_mejoras_table(ws, vivienda, catalog, header_end_row)
        _apply_a4_print_setup(ws)
    
    print_info("Populating constructor summary with dynamic formulas...")
    populate_constructor_summary(wb, catalog, viviendas)
    
    timestamp = get_timestamp()
    output_file = os.path.join(output_dir, f"viviendas_final_{timestamp}.xlsx")
    
    try:
        wb.save(output_file)
        print_success(f"Excel saved: {output_file}")
        print_info(f"Total sheets: {len(wb.sheetnames)}")
        return 0
    except Exception as e:
        print_error(f"Error saving Excel: {e}")
        return 1


if __name__ == "__main__":
    exit(generate_final_sheets())
