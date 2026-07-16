#!/usr/bin/env python3
"""
STARK096 — ALDEA: Generate Final Sheets
Creates individual sheets per housing unit from the base template
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import glob
from utils import (
    create_output_dir,
    get_timestamp,
    print_success,
    print_error,
    print_info,
    print_warning
)


def get_latest_template():
    """Get the most recent template file from output directory"""
    output_dir = "output"
    pattern = os.path.join(output_dir, "template_viviendas_*.xlsx")
    files = glob.glob(pattern)
    
    if not files:
        print_error("No template files found in output/ directory")
        return None
    
    # Get the most recent file
    latest_file = max(files, key=os.path.getmtime)
    print_info(f"Using template: {os.path.basename(latest_file)}")
    return latest_file


def read_housing_data(filepath):
    """Read housing data from the template Excel file"""
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        # Read headers (row 1)
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            headers.append(header)
        
        # Read data rows
        viviendas = []
        for row in range(2, ws.max_row + 1):
            vivienda = {}
            for col in range(1, ws.max_column + 1):
                header = headers[col - 1]
                value = ws.cell(row=row, column=col).value
                vivienda[header] = value
            viviendas.append(vivienda)
        
        wb.close()
        print_info(f"Read {len(viviendas)} housing units")
        return viviendas
    
    except Exception as e:
        print_error(f"Error reading template: {e}")
        return None


def create_header_box(ws, vivienda):
    """Create the header box in the top-left of the sheet"""
    
    # Define professional styles matching the image
    title_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    title_font = Font(bold=True, size=12, name="Calibri", color="FFFFFF")
    
    label_font = Font(bold=True, size=11, name="Calibri", color="000000")
    value_font = Font(size=11, name="Calibri", color="000000")
    
    total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    total_font = Font(bold=True, size=11, name="Calibri", color="000000")
    
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    current_row = 1
    
    # === ROW 1: FICHA DE CLIENTE ===
    ws.merge_cells(f'A{current_row}:D{current_row}')
    cell = ws.cell(row=current_row, column=1, value="FICHA DE CLIENTE")
    cell.font = title_font
    cell.fill = title_fill
    cell.border = border_style
    cell.alignment = center_align
    for col in range(2, 5):
        ws.cell(row=current_row, column=col).border = border_style
    current_row += 1
    
    # === ROW 2: NOMBRE Y APELLIDOS (COMPRADORES) ===
    # Build compradores string
    comprador1_nombre = vivienda.get("Comprador 1 - Nombre")
    comprador1_ap1 = vivienda.get("Comprador 1 - Apellido 1")
    comprador1_ap2 = vivienda.get("Comprador 1 - Apellido 2")
    
    compradores_text = ""
    if comprador1_nombre and comprador1_ap1 and comprador1_ap2:
        compradores_text = f"{comprador1_nombre} {comprador1_ap1} {comprador1_ap2}"
    
    # Check for second buyer
    comprador2_nombre = vivienda.get("Comprador 2 - Nombre")
    comprador2_ap1 = vivienda.get("Comprador 2 - Apellido 1")
    comprador2_ap2 = vivienda.get("Comprador 2 - Apellido 2")
    
    has_second_buyer = comprador2_nombre and comprador2_ap1 and comprador2_ap2
    
    # Row with first buyer
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell_label = ws.cell(row=current_row, column=1, value="Nombre y Apellidos:")
    cell_label.font = label_font
    cell_label.border = border_style
    cell_label.alignment = right_align
    ws.cell(row=current_row, column=2).border = border_style
    
    ws.merge_cells(f'C{current_row}:D{current_row}')
    cell_value = ws.cell(row=current_row, column=3, value=compradores_text)
    cell_value.font = value_font
    cell_value.border = border_style
    cell_value.alignment = left_align
    ws.cell(row=current_row, column=4).border = border_style
    current_row += 1
    
    # === ROW 3: SECOND BUYER OR EMPTY ROW ===
    if has_second_buyer:
        # Add second buyer row
        comprador2_text = f"{comprador2_nombre} {comprador2_ap1} {comprador2_ap2}"
        
        ws.merge_cells(f'A{current_row}:B{current_row}')
        cell_label = ws.cell(row=current_row, column=1, value="")
        cell_label.border = border_style
        ws.cell(row=current_row, column=2).border = border_style
        
        ws.merge_cells(f'C{current_row}:D{current_row}')
        cell_value = ws.cell(row=current_row, column=3, value=comprador2_text)
        cell_value.font = value_font
        cell_value.border = border_style
        cell_value.alignment = left_align
        ws.cell(row=current_row, column=4).border = border_style
        current_row += 1
        
        # Add separator row
        ws.merge_cells(f'A{current_row}:D{current_row}')
        cell = ws.cell(row=current_row, column=1, value="")
        cell.border = border_style
        for col in range(2, 5):
            ws.cell(row=current_row, column=col).border = border_style
        current_row += 1
    else:
        # Just empty separator row
        ws.merge_cells(f'A{current_row}:D{current_row}')
        cell = ws.cell(row=current_row, column=1, value="")
        cell.border = border_style
        for col in range(2, 5):
            ws.cell(row=current_row, column=col).border = border_style
        current_row += 1
    
    # === ROW 4/5: VIVIENDA ===
    piso = vivienda.get("Piso", "")
    # Simplify vivienda format (e.g., "Escalera 1 - Planta 4 - Puerta A" -> "E1 4A")
    vivienda_short = piso.replace("Escalera ", "E").replace(" - Planta ", " ").replace(" - Puerta ", "")
    
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell_label = ws.cell(row=current_row, column=1, value="Vivienda:")
    cell_label.font = label_font
    cell_label.border = border_style
    cell_label.alignment = right_align
    ws.cell(row=current_row, column=2).border = border_style
    
    ws.merge_cells(f'C{current_row}:D{current_row}')
    cell_value = ws.cell(row=current_row, column=3, value=vivienda_short)
    cell_value.font = value_font
    cell_value.border = border_style
    cell_value.alignment = left_align
    ws.cell(row=current_row, column=4).border = border_style
    current_row += 1
    
    # === EMPTY ROW ===
    ws.merge_cells(f'A{current_row}:D{current_row}')
    cell = ws.cell(row=current_row, column=1, value="")
    cell.border = border_style
    for col in range(2, 5):
        ws.cell(row=current_row, column=col).border = border_style
    current_row += 1
    
    # === TOTAL GASTADO ===
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell_label = ws.cell(row=current_row, column=1, value="TOTAL GASTADO:")
    cell_label.font = total_font
    cell_label.fill = total_fill
    cell_label.border = border_style
    cell_label.alignment = right_align
    ws.cell(row=current_row, column=2).border = border_style
    ws.cell(row=current_row, column=2).fill = total_fill
    
    cell_value = ws.cell(row=current_row, column=3, value="")
    cell_value.fill = total_fill
    cell_value.border = border_style
    cell_value.alignment = right_align
    
    cell_currency = ws.cell(row=current_row, column=4, value="+ IVA")
    cell_currency.font = value_font
    cell_currency.fill = total_fill
    cell_currency.border = border_style
    cell_currency.alignment = left_align
    
    # Adjust column widths for better presentation
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12


def create_summary_sheet(wb, viviendas):
    """Create the summary sheet"""
    ws = wb.active
    ws.title = "Resumen"
    
    # Header
    ws.cell(row=1, column=1, value="RESUMEN DE VIVIENDAS")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, name="Calibri")
    
    # Summary info
    ws.cell(row=3, column=1, value="Total de viviendas:")
    ws.cell(row=3, column=2, value=len(viviendas))
    
    viviendas_con_1_comprador = sum(1 for v in viviendas if not v.get("Comprador 2 - Nombre", ""))
    viviendas_con_2_compradores = len(viviendas) - viviendas_con_1_comprador
    
    ws.cell(row=4, column=1, value="Con 1 comprador:")
    ws.cell(row=4, column=2, value=viviendas_con_1_comprador)
    
    ws.cell(row=5, column=1, value="Con 2 compradores:")
    ws.cell(row=5, column=2, value=viviendas_con_2_compradores)
    
    # Style
    for row in range(3, 6):
        ws.cell(row=row, column=1).font = Font(bold=True, size=11, name="Calibri")
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15


def generate_final_sheets():
    """Generate the final Excel with individual sheets per housing unit"""
    
    print_info("Generating final Excel with individual sheets...")
    
    # Get latest template
    template_file = get_latest_template()
    if not template_file:
        return 1
    
    # Read housing data
    viviendas = read_housing_data(template_file)
    if not viviendas:
        return 1
    
    # Create output directory
    output_dir = "output_final"
    create_output_dir(output_dir)
    
    # Create new workbook
    wb = openpyxl.Workbook()
    
    # Create summary sheet
    print_info("Creating summary sheet...")
    create_summary_sheet(wb, viviendas)
    
    # Create a sheet for each vivienda
    print_info(f"Creating {len(viviendas)} individual sheets...")
    
    for idx, vivienda in enumerate(viviendas, start=1):
        piso = vivienda.get("Piso", f"Vivienda_{idx}")
        
        # Sanitize sheet name (Excel limits: max 31 chars, no special chars)
        sheet_name = piso.replace(":", "-").replace("/", "-").replace("\\", "-")
        sheet_name = sheet_name[:31]  # Max 31 characters for Excel
        
        try:
            ws = wb.create_sheet(title=sheet_name)
            create_header_box(ws, vivienda)
            
            if idx % 10 == 0:
                print_info(f"  Created {idx}/{len(viviendas)} sheets...")
        
        except Exception as e:
            print_warning(f"Could not create sheet '{sheet_name}': {e}")
            continue
    
    # Save the file
    output_file = f"{output_dir}/viviendas_final_{get_timestamp()}.xlsx"
    wb.save(output_file)
    
    print_success(f"Final Excel created: {output_file}")
    print_info(f"Total sheets: {len(wb.sheetnames)} (1 summary + {len(viviendas)} viviendas)")
    
    return 0


if __name__ == "__main__":
    try:
        exit(generate_final_sheets())
    except Exception as e:
        print_error(f"Error: {e}")
        exit(1)
