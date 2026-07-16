#!/usr/bin/env python3
"""
STARK096 — ALDEA: Excel Template Generator
Generates parametric Excel templates for construction company
"""

import openpyxl
from datetime import datetime
import os
import random
from utils import (
    load_config,
    create_output_dir,
    get_timestamp,
    apply_header_style,
    apply_cell_border,
    print_success,
    print_info
)


def generate_housing_data():
    """Generate 91 housing units data with funny dog-based names"""
    
    # Funny dog names for buyers
    nombres = [
        "Firulais", "Chucho", "Toby", "Max", "Luna", "Rocky", "Bella", "Rex",
        "Lassie", "Scooby", "Pluto", "Goofy", "Bruno", "Balto", "Laika", "Bingo",
        "Pongo", "Perla", "Boby", "Pipo", "Manchas", "Colmillo", "Pelusa", "Canelo",
        "Sultán", "Terry", "Hachiko", "Beethoven", "Marley", "Pancho", "Krypto",
        "Spike", "Snoopy", "Clifford", "Dino", "Astro", "Droopy", "Odie", "Brian",
        "Bolt", "Beethoven", "Marmaduke", "Toto", "Bruiser", "Duke", "Zeus",
        "Cooper", "Bear", "Tucker", "Oliver", "Teddy", "Murphy", "Leo", "Bentley",
        "Jack", "Lucky", "Milo", "Buddy", "Oscar", "Rusty", "Sam", "Winston",
        "Hank", "Moose", "Walter", "Buster", "Otis", "Rufus", "Jasper", "Archie",
        "Finn", "Chester", "Gus", "Louie", "Gunner", "Percy", "Scout", "Ranger",
        "Ace", "Apollo", "Bandit", "Chase", "Diesel", "Elvis", "Flash", "Ghost",
        "Hunter", "King", "Maverick", "Ninja", "Prince", "Rebel", "Storm", "Tank"
    ]
    
    apellido1_list = [
        "Ladrón", "Mordedor", "Saltarín", "Gruñón", "Olfateador", "Rascador",
        "Corretón", "Aullador", "Meneatodo", "Lameplatos", "Babosón", "Dormilón",
        "Juguetón", "Comilón", "Escarbabasura", "Perseguegatos", "Huesero", "Peludín",
        "Rabotorcido", "Orejón", "Patacorta", "Tromparruga", "Bigotudo", "Colilargo",
        "Trufarosa", "Ladridos", "Cazapelotas", "Destrozmuebles", "Sacudecola",
        "Patarrápida", "Hocicolargo", "Pelochino", "Moconariz", "Zarpadulce"
    ]
    
    apellido2_list = [
        "Guardián", "Valiente", "Peludo", "Leal", "Nervioso", "Tranquilo",
        "Inquieto", "Dormido", "Despierto", "Travieso", "Bondadoso", "Glotón",
        "Rápido", "Lento", "Ladrador", "Silencioso", "Cariñoso", "Esquivo",
        "Simpático", "Huraño", "Brincón", "Reposado", "Alegre", "Serio",
        "Juguetoncillo", "Sabio", "Torpe", "Gracioso", "Elegante", "Desaliñado"
    ]
    
    # Generate all 91 housing units
    viviendas = []
    id_counter = 1
    
    # Escalera 1: Plantas 1-7, Puertas A, B, C, D (28 viviendas)
    for planta in range(1, 8):
        for puerta in ['A', 'B', 'C', 'D']:
            piso_str = f"E1 {planta}{puerta}"
            viviendas.append({
                'id': id_counter,
                'piso': piso_str,
                'escalera': 1,
                'planta': planta,
                'puerta': puerta
            })
            id_counter += 1
    
    # Escalera 2: Plantas 1-7, Puertas A, B, C, D (28 viviendas)
    for planta in range(1, 8):
        for puerta in ['A', 'B', 'C', 'D']:
            piso_str = f"E2 {planta}{puerta}"
            viviendas.append({
                'id': id_counter,
                'piso': piso_str,
                'escalera': 2,
                'planta': planta,
                'puerta': puerta
            })
            id_counter += 1
    
    # Escalera 3: Plantas 1-7, Puertas A, B, C, D, E (35 viviendas)
    for planta in range(1, 8):
        for puerta in ['A', 'B', 'C', 'D', 'E']:
            piso_str = f"E3 {planta}{puerta}"
            viviendas.append({
                'id': id_counter,
                'piso': piso_str,
                'escalera': 3,
                'planta': planta,
                'puerta': puerta
            })
            id_counter += 1
    
    # Assign buyers (all have at least 1, about 40 have 2)
    random.seed(42)  # For reproducibility
    used_names = set()
    
    def get_unique_buyer():
        """Generate a unique buyer with dog-based name"""
        attempts = 0
        while attempts < 1000:
            nombre = random.choice(nombres)
            apellido1 = random.choice(apellido1_list)
            apellido2 = random.choice(apellido2_list)
            full_name = f"{nombre}_{apellido1}_{apellido2}"
            
            if full_name not in used_names:
                used_names.add(full_name)
                return nombre, apellido1, apellido2
            attempts += 1
        
        # If we run out of unique combinations, just use a numbered variant
        nombre = random.choice(nombres)
        apellido1 = random.choice(apellido1_list)
        apellido2 = random.choice(apellido2_list)
        return f"{nombre}{len(used_names)}", apellido1, apellido2
    
    # Decide which viviendas have 2 buyers (40 out of 91)
    viviendas_con_dos_compradores = random.sample(range(91), 40)
    
    for idx, vivienda in enumerate(viviendas):
        # First buyer (all viviendas have one)
        nombre1, apellido1_1, apellido2_1 = get_unique_buyer()
        vivienda['comprador1_nombre'] = nombre1
        vivienda['comprador1_apellido1'] = apellido1_1
        vivienda['comprador1_apellido2'] = apellido2_1
        
        # Second buyer (only some viviendas)
        if idx in viviendas_con_dos_compradores:
            nombre2, apellido1_2, apellido2_2 = get_unique_buyer()
            vivienda['comprador2_nombre'] = nombre2
            vivienda['comprador2_apellido1'] = apellido1_2
            vivienda['comprador2_apellido2'] = apellido2_2
        else:
            vivienda['comprador2_nombre'] = ""
            vivienda['comprador2_apellido1'] = ""
            vivienda['comprador2_apellido2'] = ""
    
    return viviendas


def create_template():
    """Create Excel template with 91 housing units and dog-based buyer names"""
    
    print_info("Creating Excel template with 91 viviendas...")
    
    # Load configuration
    config = load_config()
    
    # Generate housing data
    print_info("Generating housing data with funny dog names...")
    viviendas = generate_housing_data()
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compradores"
    
    # Define headers
    headers = [
        "Identificador",
        "Piso",
        "Comprador 1 - Nombre",
        "Comprador 1 - Apellido 1",
        "Comprador 1 - Apellido 2",
        "Comprador 2 - Nombre",
        "Comprador 2 - Apellido 1",
        "Comprador 2 - Apellido 2"
    ]
    
    # Write headers with style from configuration
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        apply_header_style(cell, config)
    
    # Write data for all 91 viviendas
    print_info(f"Writing {len(viviendas)} housing units to Excel...")
    
    for row_idx, vivienda in enumerate(viviendas, start=2):
        # Column 1: Identificador
        cell = ws.cell(row=row_idx, column=1)
        cell.value = vivienda['id']
        apply_cell_border(cell)
        
        # Column 2: Piso
        cell = ws.cell(row=row_idx, column=2)
        cell.value = vivienda['piso']
        apply_cell_border(cell)
        
        # Columns 3-5: Comprador 1
        cell = ws.cell(row=row_idx, column=3)
        cell.value = vivienda['comprador1_nombre']
        apply_cell_border(cell)
        
        cell = ws.cell(row=row_idx, column=4)
        cell.value = vivienda['comprador1_apellido1']
        apply_cell_border(cell)
        
        cell = ws.cell(row=row_idx, column=5)
        cell.value = vivienda['comprador1_apellido2']
        apply_cell_border(cell)
        
        # Columns 6-8: Comprador 2 (if exists)
        cell = ws.cell(row=row_idx, column=6)
        cell.value = vivienda['comprador2_nombre']
        apply_cell_border(cell)
        
        cell = ws.cell(row=row_idx, column=7)
        cell.value = vivienda['comprador2_apellido1']
        apply_cell_border(cell)
        
        cell = ws.cell(row=row_idx, column=8)
        cell.value = vivienda['comprador2_apellido2']
        apply_cell_border(cell)
    
    # Adjust column widths
    column_widths = [12, 25, 15, 18, 18, 15, 18, 18]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    
    # Freeze top row (header)
    ws.freeze_panes = 'A2'
    
    # Enable auto-filter
    ws.auto_filter.ref = ws.dimensions
    
    # Create output directory
    output_dir = create_output_dir()
    
    # Save the file with timestamp
    output_file = f"{output_dir}/template_viviendas_{get_timestamp()}.xlsx"
    wb.save(output_file)
    
    # Count viviendas with 2 buyers
    viviendas_con_dos = sum(1 for v in viviendas if v['comprador2_nombre'])
    
    print_success(f"Template created: {output_file}")
    print_info(f"Total viviendas: {len(viviendas)}")
    print_info(f"Viviendas with 1 buyer: {len(viviendas) - viviendas_con_dos}")
    print_info(f"Viviendas with 2 buyers: {viviendas_con_dos}")
    print_info(f"Escalera 1: 28 viviendas (Plantas 1-7, Puertas A-D)")
    print_info(f"Escalera 2: 28 viviendas (Plantas 1-7, Puertas A-D)")
    print_info(f"Escalera 3: 35 viviendas (Plantas 1-7, Puertas A-E)")
    
    return 0


if __name__ == "__main__":
    try:
        exit(create_template())
    except Exception as e:
        print(f"❌ Error: {e}")
        exit(1)
